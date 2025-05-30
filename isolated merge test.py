import os
import openpyxl
from openpyxl import Workbook, worksheet, load_workbook
from pathlib import Path

fpath = Path(__file__)
os.chdir(fpath.parent)
wart_234 = Workbook()


book = load_workbook(filename='WART-234.xlsx')
sheet = book.active


sheet.insert_rows(72, 1) 
sheet.move_range('A29:N29', 43)

sheet.insert_rows(64, 1)
sheet.move_range('A28:N28', 36)

sheet.insert_rows(62, 1)
sheet.move_range('A27:N27', 35)

sheet.insert_rows(58, 1) 
sheet.move_range('A26:N26', 32)

sheet.insert_rows(56, 1)
sheet.move_range('A25:N25', 31)

sheet.insert_rows(48, 1) 
sheet.move_range('A24:N24', 24)


sheet.insert_rows(46, 1)
sheet.move_range('A23:N23', 23)


sheet.insert_rows(42, 1)
sheet.move_range('A22:N22', 20)

sheet.insert_rows(38, 1) 
sheet.move_range('A21:N21', 17)

sheet.delete_rows(21, 9)


# sheet['A29'] = 'Cables & adapters'
# sheet['A34'] = 'hardware'
# sheet['A39'] ='software'
# sheet['A42'] = 'installation, programming, engineering services'
# sheet['A51'] = 'testing'
# sheet['A54'] = 'training'
# sheet['A57'] = 'other expenses'
# sheet['A62'] = 'travel'
# sheet['A71'] = 'footer'

book.save('WART-234.xlsx')