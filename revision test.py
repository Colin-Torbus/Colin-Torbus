import os
import openpyxl
from openpyxl import Workbook, worksheet, load_workbook
from pathlib import Path

fpath = Path(__file__)
os.chdir(fpath.parent)

wb = Workbook()

items = {
    'p60' : {
       'manufacturer' : 'norxe',
       'country of origin' : 'norway',
       'description' : 'cool rectangle' ,
       'price' : '$5.00',
   },
    'p10' : {
       'manufacturer' : 'norxe',
       'country of origin' : 'norway',
       'description' : 'not as cool rectangle',
       'price' : '$5.00',
       },
    'lg50' : {
       'manufacturer' : 'lg',
       'country of origin' : 'zimbabwe',
       'description' : 'crt tv',
       'price' : '$5.00',
    },
    'samsung69' : {
       'manufacturer' : 'samsung',
       'country of origin' : 'constantinople',
       'description' : '69 in tv',
       'price' : '$5.00',
   }
}

####################################################################################################

# for inputs into new_quote, add a line specifying that if the field es empty, to print 
# """TBD""" to avoid errors.
# maybe just have this be a different function for only quote info?
def new_quote(customer_id, quote_num, proj_title, job_num, proj_desc, billing_address, site_address, poc, poc_phone, poc_email):
    new_job = load_workbook(filename= '1quote template.xlsx')
    ws1 = new_job['rev1']
    ws1['H3'] = f'{customer_id}'
    ws1['H4'] = f'{quote_num}'
    ws1['H5'] = 'Rev 1'
    ws1['B6'] = f'{proj_title} \n#{job_num}'
    ws1['B10'] = f'{proj_desc}'
    ws1['G11'] = f'{billing_address}'
    ws1['J11'] = f'{site_address}'
    ws1['J14'] = f'POC: {poc}'
    ws1['J15'] = f'Phone: {poc_phone}'
    ws1['J16'] = f'Email: {poc_email}'
    new_job.title = f'{customer_id} - {job_num}'
    jobs.update(new_job.title)
    new_job.save(f'{job_num}.xlsx')
    row = 21
    line_num = 1 
    return row, line_num
#   Am i using return wrong? i cant for the life of me make it so that those variables apply to add_quote_item

####################################################################################################

#FIXME
# make the rev_num increment actually work
rev_num = 1
def quote_rev(quote_num, rev_num):
    new_rev = Workbook()
    new_rev = load_workbook(filename= f'{quote_num}.xlsx')
    rev = new_rev[f'rev{rev_num}']
    rev = new_rev.active
    n_rev = new_rev.copy_worksheet(rev)
    rev_num += 1
    n_rev.title = f'rev{rev_num}'
    new_rev.move_sheet(f'rev{rev_num}', 0)
    
#    for item in category:
#        add_quote_item(quote_num, model, qty)
#        copy formatting and insert new row
#        possibly add quote category headers through a function?
#        not sure how I would do that without breaking the bottom part of the sheet.
#    just a temp model of how we could do quote revs
#    
    
    
    new_rev.save(f'{quote_num}.xlsx')
    return rev_num
quote_rev('WART-234', rev_num)
quote_rev('WART-234', rev_num)

####################################################################################################

def add_quote_item(quote_num, model, qty):
    ##FIXME re-implement once we decide on a dict structure
    # if model not in items:
    #     print('item not found')
    # else:
    #     pass

    line_item = Workbook()
    line_item = load_workbook(filename= f'{quote_num}.xlsx')
    rev1 = line_item.active
    rev1[f'B{row}'] = line_num                                        # taken from input
    rev1[f'C{row}'] = qty                                             # taken from input
    rev1[f'D{row}'] = items [model]['manufacturer']                   ##FIXME optimize once we decide on dict sructure
    rev1[f'E{row}'] = model                                           # taken from input
    rev1[f'F{row}'] = items [model]['description']  
    rev1[f'N{row}'] = items [model]['price']
    
    line_num += 1
    row += 1
    line_item.save(f'{quote_num}.xlsx')
    return row, line_num

####################################################################################################

new_quote('WART-001', '7059', 'MITAGS Bridge 1 Upgrade', 'WART-234', 'description', 'the moon', 'antarctica',
          'Commander Erwin Smith', '5687', 'email is for nerds')
add_quote_item('WART-234', 'p60', 5)
add_quote_item('WART-234', 'lg50', 5)

