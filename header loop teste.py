import os
import openpyxl
from openpyxl import Workbook, worksheet, load_workbook
from pathlib import Path


class quote_info(Workbook):
    def __init__(self):
        self.line = 1
        self.quote_items = []
        self = Workbook()

    def print_info(self):
        print(self.quote_items)

    def add_quote_item(self, quote_num, header, category, model, qty):
        book = load_workbook(filename=f'{quote_num}.xlsx')
        self.sheet = book.active   
        row = self.line + 29
        #self.sheet.merge_cells(f'E{row}:I{row}')
        self.sheet[f'A{row}'] = self.line //2 + 1
        self.sheet[f'B{row}'] = qty
        self.sheet[f'C{row}'] = items[header][category][model]['manufacturer']
        self.sheet[f'D{row}'] = model
        self.sheet[f'E{row}'] = items[header][category][model]['description']
        self.sheet[f'M{row}'] = items[header][category][model]['price']
        book.save(f'{quote_num}.xlsx')
        self.quote_items.append(header)
        self.line += 1
        return self.line

    def quote_header(self, quote_num):
        book = load_workbook(filename=f'{quote_num}.xlsx')
        self.sheet = book.active
        
        headers = ['devices', 'cables & adapters', 'hardware', 'software',
               'installation, programming, engineering services', 'testing',
               'training','other expenses', 'travel']
        head_idx = {}
        head_offset = 0
        
        for head in headers:
            if head == 'devices':
                head_idx['devices'] = self.quote_items.count('devices') + 30
                head_offset += head_idx['devices']
            else:
                head_idx[head] = self.quote_items.count(head) * 2 + head_offset + 1
                head_offset = head_idx[head]
                
        headers = []
        for head in head_idx:
            headers.append(head_idx[head])
        head_offset = 21
        for head in headers:
            self.sheet.insert_rows(head)
            self.sheet.move_range(f'A{head_offset}:N{head_offset}', head - head_offset)
            head_offset += 1
    
        self.sheet.delete_rows(21, 9)
        book.save(f'{quote_num}.xlsx')


def new_quote(customer_id, quote_num, proj_title, job_num, proj_desc, billing_address, site_address, poc, poc_phone, poc_email):
    new_job = load_workbook(filename= 'quote template.xlsx')
    ws1 = new_job['rev1']
    ws1['G3'] = f'{customer_id}'
    ws1['G4'] = f'{quote_num}'
    ws1['G5'] = 'Rev 1'
    ws1['A6'] = f'{proj_title} \n#{job_num}'
    ws1['A10'] = f'{proj_desc}'
    ws1['F11'] = f'{billing_address}'
    ws1['I11'] = f'{site_address}'
    ws1['J14'] = f'POC: {poc}'
    ws1['J15'] = f'Phone: {poc_phone}'
    ws1['J16'] = f'Email: {poc_email}'
    new_job.title = f'{customer_id} - {job_num}'
    #  jobs.update(new_job.title)     TODO- new dict/list for jobs
    new_job.save(f'{job_num}.xlsx')


items = {
    'devices' : {
        'projector': {
            'p60' : {
                'manufacturer' : 'norxe',
                'country of origin' : 'norway',
                'description' : 'cool rectangle',
                'price' : '2.71'
            },
            'p10' : {
                'manufacturer' : 'norxe',
                'country of origin' : 'norway',
                'description' : 'not as cool rectangle',
                'price' : '5.86'
            },
        },
        'monitor' : {
            'lg50' : {
                'manufacturer' : 'lg',
                'country of origin' : 'zimbabwe',
                'description' : 'crt tv',
                'price' : '4.29'
            },
            'samsung69' : {
                'manufacturer' : 'samsung',
                'country of origin' : 'constantinople',
                'description' : '69 in tv',
                'price' : '69.69'
            },
        },
    },
    'cables & adapters' : {
        'active' : {
            'P-DPA30-50' : {
                'manufacturer' : 'Covid',
                'country of origin' : 'probably china',
                'description' : 'ACTIVE OPTICAL DISPLAYPORT CABLE, 8K, HBR3, DP 1.4, Ultra HD, with Locking Connectors, 50 ft, Plenum Rated',
                'price' : '$290.40'
            },
        },
        'passive' : {
            'NYC-604' : {
                'manufacturer' : 'New York Cables',
                'country of origin' : 'probably china',
                'description' : 'Green 1000ft Cat6 Plenum CCA Cable UTP CMP Rated 23 AWG 550 MHz Pull Box.',
                'price' : '$134.99'
            },
        },
    },
    'hardware' : {
        'mounting' : {
            'CC-106R' : {
                'manufacturer' : 'impact',
                'country of origin' : 'probably china',
                'description' : 'Super clamp for articulating camera mount.',
                'price' : '$29.95'
            },
            'BHE-107' : {
                'manufacturer' : 'impact',
                'country of origin' : 'probably china',
                'description' : '2 section articulating camera mount.',
                'price' : '$36.95'
            }  ,      
        },
    },
    'software' : {
        'scalable' : {
            'Scalable Display 11.0' : {
                'manufacturer' : 'Scalable Display Technologies',
                'country of origin' : 'United States',
                'description' : 'Software application for blending/warping geometry. Supports automatic camera based calibration system.',
                'price' : '$2650'
            },
        },
    },
    'installation, programming, engineering services' : {
        'labor' : {
            'Installation Supervisor' : {
                'manufacturer' : 'EPDS',
                'country of origin' : None,
                'description' : 'Site POC and team lead for installation. Oversees installation, configuration, alignment, testing, and training.',
                'price' : '$145',
            },
            'Installation Labor' : {
                'manufacturer' : 'EPDS',
                'country of origin' : None,
                'description' : 'On site technician for installation and configuration of equipment.',
                'price' : '$135',
            },
            'Configuration/Programming' : {
                'manufacturer' : 'EPDS',
                'country of origin' : None,
                'description' : 'EPDS technician to oversee configuration and programming of materials installed on site.',
                'price' : '$160',
            },
            'Scalable Technician' : {
                'manufacturer' : 'SDT',
                'country of origin' : None,
                'description' : 'Integrate Scalable Display into IGs and blend/warp projection system.',
                'price' : '$2250',
            },
        },
    },
    'testing' : {
        'labor' : {
            'Testing' : {
                'manufacturer' : 'EPDS',
                'country of origin' : None,
                'description' : 'Acceptance testing of visual system based on customer satisfaction/feedback. Completed during EPDS and SDT time on site.',
                'price' : '$0',
            },
        },  
    },
    'training' : {
        'labor' : {
            'Training' : {
                'manufacturer' : 'EPDS',
                'country of origin' : None,
                'description' : '## FIXME ## CUSTOM',
                'price' : '$0',
            },
        },
    },
    'other expenses' : {
        'shop' : {
            'Shop Supplies' : {
                'manufacturer' : 'EPDS',
                'country of origin' : None,
                'description' : 'Misc. fasteners, connectors, cable ties, adapters, etc. for installation.',
                'price' : '## FIXME ## CUSTOM',
            },
            'Equipment rental' : {
                'manufacturer' : 'EPDS',
                'country of origin' : None,
                'description' : '## FIXME ## CUSTOM',
                'price' : '## FIXME CUSTOM ##',
            },
        },
    },
    'travel' : {
        'transport' : {
            'Rental Car' : {
                'manufacturer' : 'EPDS',
                'country of origin' : None,
                'description' : 'Rental car for transporting EPDS/SDT personnel to and from site.',
                'price' : '## FIXME ## CUSTOM',
            },
            'Airfare' : {
                'manufacturer' : 'EPDS',
                'country of origin' : None,
                'description' : 'Round trip airfare including baggage costs. ## FIXME CUSTOM ##',
                'price' : '## FIXME ## CUSTOM',
            }
        },
        'living' : {
            'Lodging' : {
                'manufacturer' : 'EPDS',
                'country of origin' : None,
                'description' : 'Lodging for EPDS/STD personnel.',
                'price' : '## FIXME ## CUSTOM', 
            },
            'Travel Per Diem' : {
                'manufacturer' : 'EPDS',
                'country of origin' : None,
                'description' : 'Based on US GSA Per Diem Rates. Travel rates apply for EPDS/SDT staff on travel days (travel rate = standard rate x.75) ## FIXME CUSTOM ##',
                'price' : '## FIXME ## CUSTOM', 
            },
        },
    },
}


fpath = Path(__file__)
os.chdir(fpath.parent)
wart_234 = Workbook()
wart_234 = quote_info()

new_quote('WART-001', '7059', 'MITAGS Bridge 1 Upgrade', 'WART-234', 'description', 'the moon', 'antarctica',
          'Commander Erwin Smith', '5687', 'email is for nerds')
          
wart_234.add_quote_item('WART-234', 'devices', 'projector', 'p60', 5)
wart_234.add_quote_item('WART-234', 'devices', 'projector', 'p10', 2)
wart_234.add_quote_item('WART-234', 'devices', 'monitor', 'lg50', 7)
wart_234.add_quote_item('WART-234', 'devices', 'monitor', 'samsung69', 4)

wart_234.add_quote_item('WART-234', 'cables & adapters', 'active', 'P-DPA30-50', 76)
wart_234.add_quote_item('WART-234', 'cables & adapters', 'passive', 'NYC-604', 2)

wart_234.add_quote_item('WART-234', 'hardware', 'mounting', 'CC-106R', 31)
wart_234.add_quote_item('WART-234', 'hardware', 'mounting', 'BHE-107', 31)

wart_234.add_quote_item('WART-234', 'software', 'scalable', 'Scalable Display 11.0', 12)

wart_234.add_quote_item('WART-234', 'installation, programming, engineering services', 'labor', 'Installation Supervisor', 50)
wart_234.add_quote_item('WART-234', 'installation, programming, engineering services', 'labor', 'Installation Labor', 45)
wart_234.add_quote_item('WART-234', 'installation, programming, engineering services', 'labor', 'Configuration/Programming', 629)
wart_234.add_quote_item('WART-234', 'installation, programming, engineering services', 'labor', 'Scalable Technician', 2)

wart_234.add_quote_item('WART-234', 'testing', 'labor', 'Testing', 2)

wart_234.add_quote_item('WART-234', 'training', 'labor', 'Training', 5)

wart_234.add_quote_item('WART-234', 'other expenses', 'shop', 'Shop Supplies', 1)
wart_234.add_quote_item('WART-234', 'other expenses', 'shop',  'Equipment rental', 5)

wart_234.add_quote_item('WART-234', 'travel', 'transport', 'Rental Car', 2)
wart_234.add_quote_item('WART-234', 'travel', 'transport', 'Airfare', 5)
wart_234.add_quote_item('WART-234', 'travel', 'living', 'Lodging', 2)
wart_234.add_quote_item('WART-234', 'travel', 'living', 'Travel Per Diem', 58)



wart_234.quote_header('WART-234')