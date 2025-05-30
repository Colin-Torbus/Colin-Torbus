#excel format test

import os
import openpyxl
from openpyxl import Workbook, worksheet, load_workbook
from pathlib import Path

    


items = {
    'projector': {
        'p60' : {
            'manufacturer' : 'norxe',
            'country of origin' : 'norway',
            'description' : 'cool rectangle',
            'price' : '2.71'
        },
        'p10' : {
            'manufacturer' : 'norxe',
            'country of origin' : 'norway',
            'description' : 'not as cool rectangle',
            'price' : '5.86'
        }
    },
    'monitor' : {
        'lg50' : {
            'manufacturer' : 'lg',
            'country of origin' : 'zimbabwe',
            'description' : 'crt tv',
            'price' : '4.29'
        },
        'samsung69' : {
            'manufacturer' : 'samsung',
            'country of origin' : 'constantinople',
            'description' : '69 in tv',
            'price' : '69.69'
        }
    }
}


def new_item(category, manufacturer, model, origin, description, warranty, taa):
    if category in items:
        items[category][model] = {
            'manufacturer': manufacturer,
            'country of origin': origin,
            'description': description,
            'warranty' : warranty,
            'taa' : taa,
        }
    else:
        items[category] = {
            model: {
                'manufacturer': manufacturer,
                'country of origin': origin,
                'description': description,
                'warranty' : warranty,
                'taa' : taa,
            }
        }

#new_item('projector', 'norxe', 'p55', 'norway', 'weird rectangle', '5 years', 'yes')


def new_quote(customer_id, quote_num, proj_title, job_num, proj_desc, billing_address, site_address, poc, poc_phone, poc_email):
    new_job = load_workbook(filename= 'new quote template.xlsx')
    ws1 = new_job['rev1']
    ws1['G3'] = f'{customer_id}'
    ws1['G4'] = f'{quote_num}'
    ws1['G5'] = 'Rev 1'
    ws1['A6'] = f'{proj_title} \n#{job_num}'
    ws1['A10'] = f'{proj_desc}'
    ws1['F11'] = f'{billing_address}'
    ws1['I11'] = f'{site_address}'
    ws1['I14'] = f'POC: {poc}'
    ws1['I15'] = f'Phone: {poc_phone}'
    ws1['I16'] = f'Email: {poc_email}'
    ws1['A21'] = 1
    new_job.title = f'{customer_id} - {job_num}'
#    jobs.update(new_job.title)     TODO- new dict/list for jobs
    new_job.save(f'{job_num}.xlsx')


def quote_rev(quote_num, rev_num):
    revision = Workbook()
    revision = load_workbook(filename= f'{quote_num}.xlsx')
    rev1 = revision.active
    n_rev = revision.copy_worksheet(rev1)
    n_rev.title = f'rev{rev_num}'
    revision.move_sheet(f'rev{rev_num}', 0)
    revision.save(f'{quote_num}.xlsx')
    
#    changes ={ 
#        'change1' : {
#            category = input('enter category') : {
#                model = input('enter model')
#            }
#        }
#    }
#    line = 1
#    for ln_item in enumerate(changes):
#        add_quote_item(quote_num, changes[category][model], qty, line)
        







class revision():
    def __init__(self):
        self.line = 1


    def add_quote_item(self, quote_num, category, model, qty):
        book = load_workbook(filename=f'{quote_num}.xlsx')
        self.sheet = book.active   
        row = self.line + 20
        self.sheet.merge_cells(f'E{row}:I{row}')
        self.sheet[f'A{row}'] = self.line
        self.sheet[f'B{row}'] = qty
        self.sheet[f'C{row}'] = items[category][model]['manufacturer']
        self.sheet[f'D{row}'] = model
        self.sheet[f'E{row}'] = items[category][model]['description']
        self.sheet[f'M{row}'] = items[category][model]['price']
        book.save(f'{quote_num}.xlsx')
        self.line += 2
        return self.line


wart_234 = Workbook()    
wart_234 = revision()
fpath = Path(__file__)
os.chdir(fpath.parent)


wart_234.add_quote_item('WART-234', 'projector', 'p60', 2)
wart_234.add_quote_item('WART-234', 'monitor', 'lg50', 5)


#add_quote_item('WART-234', 'p10', 5)
#quote_rev('WART-234', 1)
#new_quote('WART-001', '7059', 'MITAGS Bridge 1 Upgrade', 'WART-234', 'description', 'the moon', 'antarctica',
#          'Commander Erwin Smith', '5687', 'email is for nerds')




#B21 : O24